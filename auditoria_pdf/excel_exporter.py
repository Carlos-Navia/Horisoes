from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from auditoria_pdf.domain import AuditReport, DocumentType, ParsedDocument, RuleResult


class AuditExcelExporter:
    _BASE_HEADERS = [
        "carpeta",
        "tipo_archivo",
        "documento_referencia",
        "documentos_detectados",
        "estado_documento",
        "nombre_referencia",
        "nombres_detectados",
        "estado_nombre",
        "regimen_factura",
        "regimen_detectado",
        "estado_regimen",
        "error",
    ]
    _NAME_HEADERS = {"nombre_referencia", "nombres_detectados", "estado_nombre"}

    RULE_CODE = "R1_CODIGO_FEV_VS_PDE"
    RULE_DOCUMENT = "R2_DOCUMENTO_FEV_VS_TODOS"
    RULE_REGIMEN = "R3_REGIMEN_FEV_VS_PDE"

    def __init__(self, include_name_columns: bool = True) -> None:
        self._include_name_columns = include_name_columns
        if include_name_columns:
            self.HEADERS = list(self._BASE_HEADERS)
        else:
            self.HEADERS = [
                h for h in self._BASE_HEADERS if h not in self._NAME_HEADERS
            ]

    def export(self, report: AuditReport, output_path: Path) -> Path:
        rows = self._build_rows(report)
        return self._write_rows(rows, output_path)

    def export_many(self, reports: list[tuple[str, AuditReport]], output_path: Path) -> Path:
        all_rows: list[dict[str, str]] = []
        for folder_label, report in reports:
            all_rows.extend(self._build_rows(report, folder_label=folder_label))

        if not all_rows:
            all_rows = [self._empty_row_with_error("No se detectaron carpetas con PDFs.")]

        return self._write_rows(all_rows, output_path)

    def _write_rows(self, rows: list[dict[str, str]], output_path: Path) -> Path:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "resultado"
        worksheet.append(self.HEADERS)

        for row in rows:
            worksheet.append([row.get(header, "") for header in self.HEADERS])

        self._style_sheet(worksheet)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return output_path

    def _build_rows(
        self, report: AuditReport, folder_label: str | None = None
    ) -> list[dict[str, str]]:
        parsed_documents = sorted(
            report.parsed_documents,
            key=lambda item: (item.source_path.parent.name.upper(), item.source_path.name.upper()),
        )
        if not parsed_documents:
            return [
                self._empty_row_with_error(
                    "No se detectaron documentos procesados.", folder_label=folder_label
                )
            ]

        factura = self._find_document(report, DocumentType.FACTURA)
        autorizacion = self._find_document(report, DocumentType.AUTORIZACION)

        document_reference = factura.patient_document if factura else None
        name_reference = factura.patient_name if (factura and self._include_name_columns) else None
        regimen_factura = factura.regimen if factura else None
        regimen_detectado = autorizacion.regimen if autorizacion else None

        rule_code = self._find_rule(report, self.RULE_CODE)
        rule_document = self._find_rule(report, self.RULE_DOCUMENT)
        rule_regimen = self._find_rule(report, self.RULE_REGIMEN)

        estado_regimen = self._build_regimen_status(rule_regimen)
        global_errors = "; ".join(report.errors) if report.errors else ""

        rows: list[dict[str, str]] = []
        for parsed in parsed_documents:
            estado_documento = self._build_document_status(parsed, document_reference)
            estado_nombre = (
                self._build_name_status(parsed, name_reference)
                if self._include_name_columns
                else "N/A"
            )
            error_value = self._build_row_error(
                parsed=parsed,
                estado_documento=estado_documento,
                estado_nombre=estado_nombre,
                rule_code=rule_code,
                rule_document=rule_document,
                rule_regimen=rule_regimen,
                global_errors=global_errors,
            )
            row: dict[str, str] = {
                "carpeta": folder_label or parsed.source_path.parent.name,
                "tipo_archivo": parsed.prefix,
                "documento_referencia": document_reference or "N/D",
                "documentos_detectados": parsed.patient_document or "N/D",
                "estado_documento": estado_documento,
                "regimen_factura": regimen_factura or "N/D",
                "regimen_detectado": regimen_detectado or "N/D",
                "estado_regimen": estado_regimen,
                "error": error_value,
            }
            if self._include_name_columns:
                row["nombre_referencia"] = name_reference or "N/D"
                row["nombres_detectados"] = parsed.patient_name or "N/D"
                row["estado_nombre"] = estado_nombre
            rows.append(row)
        return rows

    def _find_document(
        self,
        report: AuditReport,
        document_type: DocumentType,
    ) -> ParsedDocument | None:
        document = report.context.get_mandatory(document_type)
        if document is not None:
            return document

        for candidate in report.context.additional_documents:
            if candidate.doc_type == document_type:
                return candidate
        return None

    def _build_document_status(
        self, parsed: ParsedDocument, document_reference: str | None
    ) -> str:
        if parsed.prefix == "FEV":
            return "REFERENCIA"
        if not document_reference or not parsed.patient_document:
            return "NO_DETECTADO"
        return "COINCIDE" if parsed.patient_document == document_reference else "NO_COINCIDE"

    def _build_name_status(
        self, parsed: ParsedDocument, name_reference: str | None
    ) -> str:
        if parsed.prefix == "FEV":
            return "REFERENCIA"
        if not name_reference or not parsed.patient_name:
            return "NO_DETECTADO"
        return "COINCIDE" if parsed.patient_name == name_reference else "NO_COINCIDE"

    def _build_regimen_status(self, rule: RuleResult | None) -> str:
        if not rule:
            return "NO_EVAL"
        if rule.passed:
            return "COINCIDE"
        detail_blob = " ".join(rule.details).upper()
        if "NO SE PUDO EXTRAER" in detail_blob:
            return "NO_DETECTADO"
        return "NO_COINCIDE"

    def _build_row_error(
        self,
        parsed: ParsedDocument,
        estado_documento: str,
        estado_nombre: str,
        rule_code: RuleResult | None,
        rule_document: RuleResult | None,
        rule_regimen: RuleResult | None,
        global_errors: str,
    ) -> str:
        errors: list[str] = []
        if global_errors:
            errors.append(global_errors)

        if self._include_name_columns:
            identity_matches = (
                estado_documento in {"REFERENCIA", "COINCIDE"}
                or estado_nombre in {"REFERENCIA", "COINCIDE"}
            )
        else:
            identity_matches = estado_documento in {"REFERENCIA", "COINCIDE"}

        if not identity_matches:
            if estado_documento == "NO_DETECTADO":
                errors.append("Documento no detectado en este PDF.")
            elif estado_documento == "NO_COINCIDE":
                errors.append("Documento diferente al de referencia FEV.")

            if self._include_name_columns:
                if estado_nombre == "NO_DETECTADO":
                    errors.append("Nombre no detectado en este PDF.")
                elif estado_nombre == "NO_COINCIDE":
                    errors.append("Nombre diferente al de referencia FEV.")

        if parsed.prefix in {"FEV", "PDE"} and rule_code and not rule_code.passed:
            errors.extend(rule_code.details[:1])

        if parsed.prefix in {"FEV", "PDE"} and rule_regimen and not rule_regimen.passed:
            errors.extend(rule_regimen.details[:1])

        if rule_document and not rule_document.passed and parsed.prefix not in {"FEV", "PDE"}:
            errors.extend(rule_document.details[:1])

        if not errors:
            return ""
        deduplicated = list(dict.fromkeys(errors))
        return " | ".join(deduplicated)

    def _find_rule(self, report: AuditReport, rule_id: str) -> RuleResult | None:
        for rule in report.rule_results:
            if rule.rule_id == rule_id:
                return rule
        return None

    def _style_sheet(self, worksheet) -> None:
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(bold=True)

        for col_idx, header in enumerate(self.HEADERS, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

            max_len = len(header)
            for row_idx in range(2, worksheet.max_row + 1):
                value = worksheet.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))

            width = min(max(12, max_len + 2), 60)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = width

        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.freeze_panes = "A2"

    def _empty_row_with_error(
        self, error: str, folder_label: str | None = None
    ) -> dict[str, str]:
        row: dict[str, str] = {
            "carpeta": folder_label or "",
            "tipo_archivo": "",
            "documento_referencia": "N/D",
            "documentos_detectados": "N/D",
            "estado_documento": "NO_DETECTADO",
            "regimen_factura": "N/D",
            "regimen_detectado": "N/D",
            "estado_regimen": "NO_EVAL",
            "error": error,
        }
        if self._include_name_columns:
            row["nombre_referencia"] = "N/D"
            row["nombres_detectados"] = "N/D"
            row["estado_nombre"] = "NO_DETECTADO"
        return row
